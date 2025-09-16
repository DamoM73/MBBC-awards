import re
from pathlib import Path
import pandas as pd
from .constants import YEAR_RANGE
from .reader import read_sheet_flex
from .processor import process_year

def infer_year_from_filename(p: Path) -> int | None:
    """
    Extracts the year from the filename using regex.
    Returns the year as int if found, else None.
    """
    m = re.search(r"Year\s*(\d+)", p.stem, flags=re.IGNORECASE)
    return int(m.group(1)) if m else None

def process_file(path: Path, out_dir: Path, log_cb=print):
    """
    Processes a single Excel file:
    - Infers year from filename
    - Reads the sheet flexibly
    - Processes the year data
    - Reorders columns for output
    - Formats the output Excel sheets
    - Logs status and errors
    """
    import pandas as pd
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment

    def _reorder_award(out_df: pd.DataFrame) -> pd.DataFrame:
        """
        Reorders columns in the output DataFrame:
        - Moves 'Award' and 'Grade Point' after 'Student Name'
        """
        cols = list(out_df.columns)
        try:
            name_idx = next(i for i, c in enumerate(cols) if c.lower().startswith("student_name"))
        except StopIteration:
            name_idx = 0
        for special in ["Award", "Grade Point"]:
            if special in cols:
                cols.remove(special)
        cols.insert(min(name_idx + 1, len(cols)), "Award")
        award_idx = cols.index("Award")
        cols.insert(award_idx + 1, "Grade Point")
        return out_df.reindex(columns=[c for c in cols if c in out_df.columns])

    def _format_ws(ws, name_col_letter=None):
        """
        Formats the worksheet:
        - Sets column widths based on content
        - Centers headers
        - Left-aligns 'Student Name' column, centers others
        """
        max_row, max_col = ws.max_row, ws.max_column
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = "" if cell.value is None else str(cell.value)
                if len(val) > max_len:
                    max_len = len(val)
                if row_idx == 1:
                    # headers always centred
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    if col_letter == name_col_letter:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
            width = min(max(10, int(max_len * 1.2)), 60)
            ws.column_dimensions[col_letter].width = width

    try:
        year = infer_year_from_filename(path)
        if year is None:
            log_cb(f"[SKIP] {path.name}: could not infer year from filename")
            return None
        if year not in YEAR_RANGE:
            log_cb(f"[SKIP] {path.name}: year {year} not in 7–10")
            return None

        # Read the Excel sheet flexibly
        df = read_sheet_flex(path)
        # Process the year data, returns awards and subject averages
        out, subj_df = process_year(df, year)

        # Reorder columns for output
        out = _reorder_award(out)

        # Prepend Student Name to subject averages DataFrame
        name_cols = [c for c in out.columns if c.lower().startswith("student_name")]
        if name_cols:
            subj_df = pd.concat([out[name_cols[0]].reset_index(drop=True),
                                 subj_df.reset_index(drop=True)], axis=1)
            subj_df.rename(columns={name_cols[0]: "Student Name"}, inplace=True)

        # Write results to Excel with formatting
        out_path = out_dir / f"{path.stem} - Awards.xlsx"
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            out.to_excel(writer, index=False, sheet_name="Raw+Awards")
            subj_df.to_excel(writer, index=False, sheet_name="Subject_Averages")
            wb = writer.book
            # Raw+Awards: student name is column B
            _format_ws(wb["Raw+Awards"], name_col_letter="B")
            # Subject_Averages: student name is column A
            _format_ws(wb["Subject_Averages"], name_col_letter="A")

        log_cb(f"[OK]   {path.name} → {out_path.name}")
        return out_path
    except Exception as e:
        log_cb(f"[ERR]  {path.name}: {e}")
        return None
